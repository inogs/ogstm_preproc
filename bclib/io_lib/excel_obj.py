import numpy as np
from openpyxl import load_workbook
class xlsx:

    """
        Class xls

    """

    def __init__(self, xls_file):
        try:
            self.wb = load_workbook(filename=xls_file, read_only=False,data_only=True)
            self.path = xls_file
        except:
            print("RIVER XLS FILE NOT FOUND")
            exit()


    def read_spreadsheet_allrow(self,worsheet,y_range):

        ws = self.wb[worsheet]
        n = len(ws.columns[1])
        a = np.zeros((n,len(y_range)))

        for x in range(2,n):
            for y in y_range:
                a[x-2][y-y_range[0]] = ws.cell(row = x , column = y).value

        return a

    def read_spreadsheet_range(self,worsheet,x_range,y_range,dtype="d"):

        ws = self.wb[worsheet]
        if dtype=="d":
            a = np.zeros((len(x_range),len(y_range)),dtype=np.float)
        if dtype=="i":
            a = np.zeros((len(x_range),len(y_range)),dtype=np.int)

        for x in x_range:
            for y in y_range:
                a[x-x_range[0]][y-y_range[0]] = ws.cell(row = x , column = y).value

        return a
