import numpy as np
from openpyxl import load_workbook
class xlsx:

    """
        Class xls

    """

    def __init__(self, xls_file):
        try:
            self.wb = load_workbook(filename=xls_file, read_only=False,data_only=True)
            self.path = xls_file
        except:
            print("RIVER XLS FILE NOT FOUND")
            exit()


    def read_spreadsheet_all(self,worsheet):

        ws = self.wb[worsheet]
        
       
        b = []
        x = 0
        while True :
            x = x+1
            if  ws.cell(row = x , column = 1).value == -999:
                    break
            
            a = []
            y = 0
            while True:
                y = y +1
                if  ws.cell(row = x , column = y).value == -999:
                    break
                a.append( ws.cell(row = x , column = y).value )
            b.append(a[:])

        cast_nparray = np.array(b)            
        return cast_nparray

    def read_spreadsheet_range(self,worsheet,x_range,y_range,dtype="d"):

        ws = self.wb[worsheet]
        if dtype=="d":
            a = np.zeros((len(x_range),len(y_range)),dtype=float)
        if dtype=="i":
            a = np.zeros((len(x_range),len(y_range)),dtype=int)

        for x in x_range:
            for y in y_range:
                a[x-x_range[0]][y-y_range[0]] = ws.cell(row = x , column = y).value

        return a
